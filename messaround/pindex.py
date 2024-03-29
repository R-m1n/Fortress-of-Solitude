import os
import random
import string
import csv
import sys
from typing import Any
import openpyxl
import time
from pathlib import Path
from alive_progress import alive_bar


class Product:
    def __init__(self, name: str, price: str,
                 discount: str, weight: str,
                 length: str, width: str,
                 height: str, categories: str) -> None:

        self.id = generate_id()
        self.type = "simple"
        self.sku = ""
        self.name = name
        self.published = 1
        self.is_featured = 0
        self.visibility = "visible"
        self.short_desc = ""
        self.desc = ""
        self.in_stock = 1
        self.weight = weight
        self.length = length
        self.width = width
        self.height = height
        self.costumer_review = 1
        self.price = price
        self.sale_price = str(
            float(price) - (float(price) * (float(discount) / 100))
        )
        self.categories = " > ".join(categories.split("-"))

    def info(self) -> list[str]:
        return [
            self.id,
            self.type,
            self.sku,
            self.name,
            self.published,
            self.is_featured,
            self.visibility,
            self.short_desc,
            self.desc,
            self.in_stock,
            self.weight,
            self.length,
            self.width,
            self.height,
            self.costumer_review,
            self.sale_price,
            self.price,
            self.categories,
        ]


def generate_id() -> str:
    id = random.choices(string.digits, k=8)

    id = "".join(id)
    if id not in id_list:
        id_list.append(id)

        return id

    generate_id()


def progress(bar: Any, text: str) -> None:
    for i in range(BARS_PER_CALL):
        bar.text(text)
        time.sleep(.04)
        bar()

    print("Done.")


if __name__ == "__main__":
    TOTAL_BAR = 100
    TOTAL_PROGRESS_CALL = 5
    BARS_PER_CALL = TOTAL_BAR // TOTAL_PROGRESS_CALL

    with alive_bar(TOTAL_BAR, force_tty=True, dual_line=True, bar='blocks') as bar:
        progress(bar, " -> Collecting source files...")

        curr_dir_files = {Path(file.name).suffix: Path(file.name) for file in os.scandir(os.getcwd())
                          if Path(file.name).suffix == ".xlsx" or Path(file.name).name == "id_record.csv"}

        if ".xlsx" not in curr_dir_files:
            print(".xlsx file not found :(")
            sys.exit()

        progress(bar, " -> Initializing ID database...")

        if ".csv" in curr_dir_files:
            with open(curr_dir_files.get(".csv"), "r") as id_record:
                reader = csv.reader(id_record)

                for row in reader:
                    id_list = row

        else:
            id_list = []

        progress(bar, " -> Preparing .xlsx file...")

        csv_file_path = Path("product_record.csv")

        source_file_path = curr_dir_files.get(".xlsx")

        xlsx_file = openpyxl.load_workbook(source_file_path)
        sheet = xlsx_file.active

        progress(bar, " -> Cooking up .csv file...")

        shelf = []
        for row in range(1, sheet.max_row + 1):
            product_info = []

            for column in range(1, sheet.max_column + 1):
                product_info.append(sheet.cell(row=row, column=column).value)

            product = Product(name=product_info[0],
                              price=product_info[1],
                              discount=product_info[2],
                              weight=product_info[3],
                              length=product_info[4],
                              width=product_info[5],
                              height=product_info[6],
                              categories=product_info[7],
                              )

            shelf.append(product.info())

        with open(csv_file_path, "w", newline='') as csv_file:
            writer = csv.writer(csv_file)

            columns = [
                "ID",
                "Type",
                "SKU",
                "Name",
                "Published",
                "is featured?",
                "Visibility in catalog",
                "Short description",
                "Description",
                "In stock?",
                "Weight",
                "Length",
                "Width",
                "Height",
                "Allow customer reviews?",
                "Sale price",
                "Regular price",
                "Categories",
            ]

            writer.writerow(columns)

            writer.writerows(shelf)

        progress(bar, " -> Saving .csv file...")
